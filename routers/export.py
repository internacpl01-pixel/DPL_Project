"""Export router — CSV, Excel (.xlsx), and PDF download of master table data."""
from __future__ import annotations

import csv
import io
import logging
from datetime import date, datetime
from typing import Optional

from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse

from fpdf import FPDF

from dependencies import get_current_user
from services.data import get_master_rows

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api", tags=["export"])

_NUMERIC_TYPES = {"numeric", "real", "double precision", "integer", "bigint"}
_INTEGER_TYPES = {"integer", "bigint", "smallint"}
_DATE_TYPES = {"date", "timestamp", "timestamp without time zone",
               "timestamp with time zone"}


# ── helpers ──────────────────────────────────────────────────────────

def _is_numeric(col_type: str) -> bool:
    return (col_type or "").lower() in _NUMERIC_TYPES


def _is_integer(col_type: str) -> bool:
    """Whole-number column — money formatting would render an id as "41.00"."""
    return (col_type or "").lower() in _INTEGER_TYPES


def _is_date(col_type: str) -> bool:
    return (col_type or "").lower() in _DATE_TYPES


def _as_date(val):
    """Coerce a cell value to date/datetime for Excel, or None if it isn't one."""
    if isinstance(val, (datetime, date)):
        return val
    s = str(val).strip()
    for parse in (date.fromisoformat, datetime.fromisoformat):
        try:
            return parse(s)
        except (ValueError, TypeError):
            continue
    return None


async def _fetch_all(search: str = "") -> tuple[list[dict], list[dict]]:
    """Return (columns, rows) with the full master table.

    Columns come from the first page; rows are paginated in 500-row chunks.
    Client-side search filtering is applied when *search* is non-empty.
    """
    limit = 500
    offset = 0
    all_rows: list[dict] = []
    columns: list[dict] = []

    while True:
        result = await get_master_rows(limit=limit, offset=offset)
        if not columns and result.get("columns"):
            columns = result["columns"]
        batch = result.get("rows", [])

        if search:
            term = search.lower()
            batch = [
                r for r in batch
                if any(str(v).lower().find(term) != -1 for v in r.values() if v is not None)
            ]

        all_rows.extend(batch)
        if len(batch) < limit:
            break
        offset += limit

    return columns, all_rows


# ── CSV ──────────────────────────────────────────────────────────────

def _build_csv(
    col_names: list[str], col_display: list[str], col_types: dict[str, str], rows: list[dict]
) -> tuple[bytes, str, str]:
    buf = io.StringIO(newline="")
    writer = csv.writer(buf, quoting=csv.QUOTE_MINIMAL)

    writer.writerow(col_display)

    for row in rows:
        out: list[str] = []
        for c in col_names:
            val = row.get(c, "")
            if val is None:
                val = ""
            elif _is_numeric(col_types.get(c, "")):
                val = str(val)
            else:
                val = str(val)
            out.append(val)
        writer.writerow(out)

    return buf.getvalue().encode("utf-8-sig"), "text/csv", "csv"


# ── Excel ────────────────────────────────────────────────────────────

def _build_xlsx(
    col_names: list[str], col_display: list[str], col_types: dict[str, str], rows: list[dict]
) -> tuple[bytes, str, str]:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="Excel export requires openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = "Master Data"

    _HDR_FILL = PatternFill(fill_type="solid", fgColor="DDDDDD")

    # header row
    for ci, label in enumerate(col_display, 1):
        cell = ws.cell(row=1, column=ci, value=label)
        cell.font = cell.font.copy(bold=True)
        cell.fill = _HDR_FILL

    # Data rows. Cells are typed from the live column type so Excel gets real
    # dates and integers instead of text — sorting, date filters and number
    # formatting all depend on the cell type, not just how the value looks.
    # Widths are measured from the RENDERED text, not the raw value: "#,##0.00"
    # turns 4000000.0 into "4,000,000.00" — three characters wider than str()
    # reports — and a column sized from str() renders the cell as "#####".
    widths = [len(str(lbl)) for lbl in col_display]

    for ri, row in enumerate(rows, 2):
        for ci, c in enumerate(col_names, 1):
            val = row.get(c, "")
            ctype = col_types.get(c, "")
            cell = ws.cell(row=ri, column=ci)
            shown = ""
            if val is None or val == "":
                cell.value = None
            elif _is_date(ctype):
                parsed = _as_date(val)
                if parsed is None:
                    cell.value = shown = str(val)
                else:
                    cell.value = parsed
                    cell.number_format = "yyyy-mm-dd"
                    shown = "0000-00-00"
            elif _is_integer(ctype):
                s = str(val).replace(",", "").strip()
                try:
                    cell.value = int(s) if s.lstrip("-").isdigit() else int(float(s))
                    cell.number_format = "0"
                    shown = str(cell.value)
                except (ValueError, TypeError):
                    cell.value = shown = str(val)
            elif _is_numeric(ctype):
                try:
                    cell.value = float(str(val).replace(",", ""))
                    cell.number_format = "#,##0.00"
                    shown = f"{cell.value:,.2f}"
                except (ValueError, TypeError):
                    cell.value = shown = str(val)
            else:
                cell.value = shown = str(val)
            widths[ci - 1] = max(widths[ci - 1], len(shown))

    # Pad by 2 for cell margins; cap at 50 so one long narration can't stretch
    # the sheet off-screen.
    for ci in range(1, len(col_names) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = min(widths[ci - 1] + 2, 50)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "xlsx"


# ── PDF ──────────────────────────────────────────────────────────────

class _ExportPDF(FPDF):
    """Table-based PDF export — fixed row heights, no wrapping chaos."""

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._title_rows = 0

    def header(self) -> None:
        if self.page_no() == 1:
            self._title_rows = 0  # first page uses regular header
        self.set_font("Helvetica", "B", 13)
        self.cell(0, 7, "Master Data Export", new_x="LMARGIN", new_y="NEXT")
        self.set_font("Helvetica", "", 9)
        self.cell(0, 5,
            f"Generated: {date.today().isoformat()}     "
            f"Rows: {self.row_count}",
            new_x="LMARGIN", new_y="NEXT")
        self.ln(1)
        self._title_rows += 2 if self.page_no() == 1 else 1

    def footer(self) -> None:
        self.set_y(-10)
        self.set_font("Helvetica", "I", 8)
        self.cell(0, 10, f"Page {self.page_no()}", align="C")


def _build_pdf(
    col_names: list[str], col_display: list[str], col_types: dict[str, str], rows: list[dict]
) -> tuple[bytes, str, str]:
    pdf = _ExportPDF(orientation="L", unit="mm", format="A4")
    pdf.row_count = len(rows)
    # Page breaks are handled manually so every page can repeat the header row;
    # leaving fpdf's automatic break on as well would split a row across pages
    # and re-enter header() mid-row with a different font size.
    pdf.set_auto_page_break(auto=False, margin=12)

    usable_w = pdf.w - pdf.l_margin - pdf.r_margin  # landscape A4 ≈ 277 mm
    n = len(col_names)

    # Widths come from the actual content, never from a name-keyed table — the
    # field map is dynamic, so any hardcoded lookup silently degrades to an even
    # split for renamed or custom fields. Weight each column by its widest
    # sampled cell, clamp so one long narration can't starve the rest, then
    # normalize so the table ends exactly on the right margin.
    sample = rows[:200]
    weights: list[float] = []
    for i, c in enumerate(col_names):
        longest = len(str(col_display[i]))
        for r in sample:
            v = r.get(c)
            if v is not None:
                longest = max(longest, len(str(v)))
        weights.append(float(min(max(longest, 6), 70)))

    total_wt = sum(weights) or 1.0
    col_w = [max(usable_w * (w / total_wt), 8.0) for w in weights]
    scale = usable_w / sum(col_w)  # the 8 mm floor can push the sum over
    col_w = [w * scale for w in col_w]

    row_h = 5.5    # header row / minimum data row height
    line_h = 3.2   # one wrapped line of body text
    pad = 0.5      # horizontal breathing room inside each cell

    def _fit(text: str, w: float) -> str:
        """Trim *text* to what fits on one line of *w* mm at the current font."""
        avail = w - 2 * pad
        if avail <= 0:
            return ""
        if pdf.get_string_width(text) <= avail:
            return text
        lo, hi = 0, len(text)
        while lo < hi:
            mid = (lo + hi + 1) // 2
            if pdf.get_string_width(text[:mid] + "...") <= avail:
                lo = mid
            else:
                hi = mid - 1
        return text[:lo] + "..." if lo else ""

    def _wrap(text: str, w: float) -> list[str]:
        """Split *text* into the lines fpdf will lay out in a *w* mm cell."""
        if not text:
            return [""]
        # WORD wrapping keeps the text reconstructable from the PDF text layer;
        # fpdf still hard-splits any single token too long for the column.
        return list(pdf.multi_cell(w - 2 * pad, line_h, text, dry_run=True,
                                   output="LINES")) or [""]

    def _header_row() -> None:
        # Must set its own font: header() leaves it at 13/9 pt, which clipped
        # the labels on the first page only.
        pdf.set_font("Helvetica", "B", 7)
        pdf.set_fill_color(230, 230, 230)
        for label, w in zip(col_display, col_w):
            pdf.cell(w, row_h, _fit(str(label), w), border=1, align="L", fill=True)
        pdf.ln(row_h)

    def _draw_row(row: dict) -> None:
        """Render one transaction, wrapping long text instead of dropping it."""
        pdf.set_font("Helvetica", "", 7)
        cells = []
        for c, w in zip(col_names, col_w):
            val = row.get(c, "")
            val = "" if val is None else str(val)
            cells.append(_wrap(val, w))

        h = max(row_h, max(len(ls) for ls in cells) * line_h + 1.4)

        if pdf.get_y() + h > pdf.h - pdf.b_margin:
            pdf.add_page()
            _header_row()
            pdf.set_font("Helvetica", "", 7)

        x, y = pdf.l_margin, pdf.get_y()
        for c, w, lines in zip(col_names, col_w, cells):
            pdf.rect(x, y, w, h)
            pdf.set_xy(x + pad, y + 0.7)
            pdf.multi_cell(w - 2 * pad, line_h, "\n".join(lines),
                           border=0, align="R" if _is_numeric(col_types.get(c, "")) else "L")
            x += w
        pdf.set_xy(pdf.l_margin, y + h)

    if n == 0 or not rows:
        pdf.add_page()
        pdf.set_font("Helvetica", "", 10)
        pdf.cell(0, 8, "No data to export.", new_x="LMARGIN", new_y="NEXT")
    else:
        pdf.add_page()
        _header_row()
        for row in rows:
            _draw_row(row)

    buf = io.BytesIO()
    pdf.output(buf)
    return buf.getvalue(), "application/pdf", "pdf"


# ── endpoint ─────────────────────────────────────────────────────────

@router.get("/export")
async def export_data(
    format: str = Query("csv", pattern="^(csv|xlsx|pdf)$"),
    search: str = Query("", description="Filter rows by text (matches any column)"),
    current_user: dict = Depends(get_current_user),
):
    columns, rows = await _fetch_all(search=search)

    if not columns:
        columns = [{"name": "id", "displayname": "ID", "type": "integer"}]

    col_names = [c["name"] for c in columns]
    col_display = [c.get("displayname") or c["name"] for c in columns]
    col_types = {c["name"]: c.get("type", "") for c in columns}

    today_str = date.today().isoformat()
    filename = f"master_data_{today_str}"

    try:
        if format == "csv":
            content, media_type, ext = _build_csv(col_names, col_display, col_types, rows)
        elif format == "xlsx":
            content, media_type, ext = _build_xlsx(col_names, col_display, col_types, rows)
        else:
            content, media_type, ext = _build_pdf(col_names, col_display, col_types, rows)
    except Exception as exc:
        logger.error(f"Export failed ({format}): {exc}")
        raise HTTPException(status_code=500, detail=f"Export failed: {exc}")

    return StreamingResponse(
        io.BytesIO(content),
        media_type=media_type,
        headers={
            "Content-Disposition": f'attachment; filename="{filename}.{ext}"',
        },
    )
